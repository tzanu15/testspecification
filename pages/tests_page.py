import copy
import json
import os
import sys
import traceback

import pandas as pd
from PyQt5.QtWidgets import (
	QWidget, QVBoxLayout, QPushButton, QTableWidget, QTableWidgetItem,
	QHBoxLayout, QLabel, QLineEdit, QHeaderView, QFrame, QInputDialog, QFileDialog, QMessageBox,
	QDialog, QListWidget, QTextEdit, QToolButton, QStyle, QTabWidget, QMenu, QAction, QListWidgetItem
)
from PyQt5.QtGui import QFont, QColor
from PyQt5.QtCore import Qt, pyqtSignal, QTimer
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Font


# ----------------- Dialog Classes -----------------
class SelectCommandDialog(QDialog):
	"""
	Dialog pentru alegerea unei comenzi din generic_commands.json.
	Utilizatorul selectează o comandă prin dublu click.
	"""

	def __init__(self, commands_data):
		super().__init__()
		self.setWindowTitle("Select a Command")
		self.setGeometry(400, 300, 400, 300)
		self.selected_command = None
		layout = QVBoxLayout()
		self.command_list = QListWidget()
		# Adaugă toate cheile din commands_data
		self.command_list.addItems(list(commands_data.keys()))
		self.command_list.itemDoubleClicked.connect(self.select_command)
		layout.addWidget(self.command_list)
		self.setLayout(layout)

	def select_command(self, item):
		self.selected_command = item.text()
		self.accept()

class SelectParameterDialog(QDialog):
	"""
	Dialog pentru alegerea parametrilor necesari pentru o comandă.
	Se afișează câte o listă pentru fiecare categorie.
	"""

	def __init__(self, parameters_data, required_categories):
		super().__init__()
		self.setWindowTitle("Select Parameters")
		self.setGeometry(400, 300, 400, 300)
		self.selected_parameters = {}
		layout = QVBoxLayout()
		self.parameter_lists = {}
		for category in required_categories:
			layout.addWidget(QLabel(f"Select {category}:"))
			param_list = QListWidget()
			param_list.addItems(parameters_data.get(category, []))
			param_list.itemDoubleClicked.connect(lambda item, cat=category: self.select_parameter(cat, item.text()))
			layout.addWidget(param_list)
			self.parameter_lists[category] = param_list
		self.setLayout(layout)

	def select_parameter(self, category, value):
		self.selected_parameters[category] = value
		if len(self.selected_parameters) == len(self.parameter_lists):
			self.accept()


from PyQt5.QtWidgets import (
	QDialog, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
	QTabWidget, QWidget, QLineEdit, QTableWidget, QTableWidgetItem
)
from PyQt5.QtCore import Qt


class PreviewTestStepDialog(QDialog):
    """Dialog pentru vizualizarea test step-ului și selectarea parametrilor."""

    def __init__(self, command_action, command_expected, parameters_data, required_categories):
        super().__init__()
        self.setWindowTitle("Preview Test Step")
        self.setGeometry(400, 300, 600, 500)
        self.selected_parameters = {}

        layout = QVBoxLayout()

        # 🔹 Action
        self.action_label = QLabel("<b>Action:</b>")
        layout.addWidget(self.action_label)

        self.action_text = QTextEdit(self.highlight_placeholders(command_action))
        self.action_text.setReadOnly(True)
        layout.addWidget(self.action_text)

        # 🔹 Expected Result
        self.expected_label = QLabel("<b>Expected:</b>")
        layout.addWidget(self.expected_label)

        self.expected_text = QTextEdit(self.highlight_placeholders(command_expected))
        self.expected_text.setReadOnly(True)
        layout.addWidget(self.expected_text)

        # 🔹 Tabs pentru parametri
        self.tab_widget = QTabWidget()
        self.parameter_lists = {}

        for category in required_categories:
            tab = QWidget()
            tab_layout = QVBoxLayout()

            # 🔹 Căutare în fiecare tab
            search_input = QLineEdit()
            search_input.setPlaceholderText(f"Search in {category}...")
            search_input.textChanged.connect(lambda text, cat=category: self.filter_parameters(cat, text))
            tab_layout.addWidget(search_input)

            param_list = QListWidget()
            param_list.addItems(parameters_data.get(category, []))
            param_list.itemDoubleClicked.connect(lambda item, cat=category: self.select_parameter(cat, item.text()))

            tab_layout.addWidget(param_list)
            tab.setLayout(tab_layout)

            self.parameter_lists[category] = param_list
            self.tab_widget.addTab(tab, category)

        layout.addWidget(self.tab_widget)

        # 🔹 Confirm Button
        self.confirm_button = QPushButton("Apply Parameters")
        self.confirm_button.setEnabled(False)
        self.confirm_button.clicked.connect(self.accept)
        layout.addWidget(self.confirm_button)

        self.setLayout(layout)

    def highlight_placeholders(self, text):
        """Evidențiază placeholder-urile în text."""
        highlighted_text = text
        for placeholder in set(part.strip("{}") for part in text.split() if "{" in part):
            highlighted_text = highlighted_text.replace(
                f"{{{placeholder}}}", f"<span style='background-color: yellow; font-weight: bold;'>{'{'+placeholder+'}'}</span>"
            )
        return f"<html><body>{highlighted_text}</body></html>"

    def select_parameter(self, category, value):
        """Salvează parametrul selectat și actualizează vizual textul fără a închide dialogul imediat."""
        self.selected_parameters[category] = value
        self.update_text_display(category, value)

    def update_text_display(self, category, value):
        """Actualizează vizual Action și Expected Result cu parametrii selectați."""
        self.action_text.setHtml(self.action_text.toHtml().replace(f"{{{category}}}", value))
        self.expected_text.setHtml(self.expected_text.toHtml().replace(f"{{{category}}}", value))

        if len(self.selected_parameters) == len(self.parameter_lists):
            self.confirm_button.setEnabled(True)

    def filter_parameters(self, category, search_text):
        """Filtrăm parametrii în tab-ul selectat pe baza textului introdus."""
        param_list = self.parameter_lists.get(category)
        if param_list:
            for i in range(param_list.count()):
                item = param_list.item(i)
                item.setHidden(search_text.lower() not in item.text().lower())


class TestStepTable(QWidget):
    step_selected = pyqtSignal(int)  # 🔹 Semnal pentru selecția unui step

    def __init__(self, steps, parent_table, row_index, linked_table=None):
        super().__init__()
        self.parent_table = parent_table  # Referință la tabelul principal
        self.row_index = row_index  # Rândul testului părinte
        self.linked_table = linked_table  # Referință la tabelul pereche (Expected <-> Action)

        layout = QVBoxLayout()
        self.table = QTableWidget()
        self.table.setColumnCount(1)
        self.table.setRowCount(len(steps))
        self.table.horizontalHeader().setVisible(False)
        self.table.setShowGrid(False)
        self.table.verticalHeader().setVisible(False)
        self.table.horizontalHeader().setStretchLastSection(True)

        self.table.setSelectionBehavior(QTableWidget.SelectRows)  # ✅ Forțăm selecția rândului
        self.table.setSelectionMode(QTableWidget.SingleSelection)  # ✅ Doar o singură selecție

        for i, step in enumerate(steps):
            item = QTableWidgetItem(step)
            item.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable)
            self.table.setItem(i, 0, item)

        # 🔹 Conectăm selecția la metoda `sync_selection`
        self.table.itemSelectionChanged.connect(self.sync_selection)

        layout.addWidget(self.table)
        self.setLayout(layout)

    def sync_selection(self):
        """Selectează automat testul părinte în tabelul principal când un step este selectat."""
        selected_row = self.table.currentRow()
        if selected_row != -1:
            print(f"✅ DEBUG: Step selected in embedded table - Index: {selected_row}")

            # 🔹 Selectăm întregul rând în tabelul principal (self.test_table)
            self.parent_table.selectRow(self.row_index)

            # 🔹 Dacă tabelul pereche există, selectăm și în el
            if self.linked_table:
                self.linked_table.table.selectRow(selected_row)

# ----------------- Main TestsPage Class -----------------
class TestsPage(QWidget):
	def __init__(self):
		super().__init__()
		self.copied_step = None
		self.copied_expected = None


		self.json_file = self.get_resource_path("../data/tests.json")
		self.commands_file = self.get_resource_path("../data/generic_commands.json")
		self.parameters_file = self.get_resource_path("../data/parameters.json")

		# 🔹 Încărcăm datele necesare
		self.tests_data = self.load_json(self.json_file)
		self.commands_data = self.load_json(self.commands_file)
		self.parameters_data = self.load_json(self.parameters_file)



		layout = QVBoxLayout()

		title = QLabel("Manage Tests")
		title.setFont(QFont("Arial", 16, QFont.Bold))
		title.setAlignment(Qt.AlignCenter)
		layout.addWidget(title)

		line = QFrame()
		line.setFrameShape(QFrame.HLine)
		line.setFrameShadow(QFrame.Sunken)
		layout.addWidget(line)

		input_layout = QHBoxLayout()

		self.test_name_input = QLineEdit()
		self.test_name_input.setPlaceholderText("Enter Test Name...")
		input_layout.addWidget(self.test_name_input)

		self.add_test_button = QPushButton("Add Test")
		self.add_test_button.clicked.connect(self.add_test)
		input_layout.addWidget(self.add_test_button)
		layout.addLayout(input_layout)

		# ✅ Search Bar for Tests
		search_layout = QHBoxLayout()
		search_bar = QLineEdit()
		search_bar.setPlaceholderText("🔍 Search tests...")
		search_bar.textChanged.connect(lambda text: self.filter_tests(text, self.test_table))  # Connect search function
		search_layout.addWidget(search_bar)
		layout.addLayout(search_layout)

		# Butoane pentru Import/Export
		button_layout = QHBoxLayout()
		self.import_button = QPushButton("Import from XLSX")
		self.import_button.clicked.connect(self.import_from_xlsx)
		button_layout.addWidget(self.import_button)

		self.export_button = QPushButton("Export to XLSX")
		self.export_button.clicked.connect(self.export_to_xlsx)
		button_layout.addWidget(self.export_button)

		layout.addLayout(button_layout)

		# 🔹 Tabel pentru afișarea testelor
		self.test_table = QTableWidget()
		self.test_table.setShowGrid(False)
		self.test_table.setWordWrap(True)
		self.test_table.setTextElideMode(Qt.ElideNone)


		self.test_table.setColumnCount(7)
		self.test_table.setHorizontalHeaderLabels([
			"Test Name", "Description", "Precondition", "Action", "Expected Results",
			"Test Data Description", "Description TCG"
		])

		self.test_table.verticalHeader().hide()
		self.test_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
		self.test_table.setContextMenuPolicy(Qt.CustomContextMenu)
		self.test_table.setSelectionBehavior(QTableWidget.SelectRows)
		self.test_table.setSelectionMode(QTableWidget.SingleSelection)
		self.test_table.customContextMenuRequested.connect(self.show_context_menu)
		self.test_table.customContextMenuRequested.connect(self.show_test_context_menu)
		self.test_table.cellChanged.connect(self.save_edited_test)
		layout.addWidget(self.test_table)
		self.setLayout(layout)

		self.load_tests()

	def add_test(self):
		"""Adaugă un test nou în tabel și în JSON."""
		test_name = self.test_name_input.text().strip()
		if not test_name:
			return

		description, ok = QInputDialog.getText(self, "Enter Description", f"Enter description for {test_name}:")
		if not ok:
			description = ""

		precondition, ok = QInputDialog.getText(self, "Enter Precondition", f"Enter precondition for {test_name}:")
		if not ok:
			precondition = ""

		row_position = self.test_table.rowCount()
		self.test_table.insertRow(row_position)

		self.test_table.setItem(row_position, 0, QTableWidgetItem(test_name))
		self.test_table.setItem(row_position, 1, QTableWidgetItem(description))
		self.test_table.setItem(row_position, 2, QTableWidgetItem(precondition))

		for col in range(3, 7):
			self.test_table.setItem(row_position, col, QTableWidgetItem(""))

		self.tests_data[test_name] = {
			"Description": description,
			"Precondition": precondition,
			"Action": "",
			"Expected Results": "",
			"Test Data Description": "",
			"Description TCG": ""
		}

		self.save_tests()
		self.test_name_input.clear()

	def add_test_step(self):
		"""Adaugă un nou Test Step la testul selectat și îl salvează corect."""
		selected_row = self.test_table.currentRow()
		if selected_row == -1:
			QMessageBox.warning(self, "No Test Selected", "Please select a test to add a step.")
			return

		test_name = self.test_table.item(selected_row, 0).text().strip()
		print(f"🔹 Selected test: {test_name}")  # Debugging

		self.load_parameters()
		self.load_commands()

		# 🔹 Selectăm comanda
		command_dialog = SelectCommandDialog(self.commands_data)
		if not command_dialog.exec_():
			print("❌ No command selected. Exiting...")
			return

		selected_command = command_dialog.selected_command
		command_action = self.commands_data[selected_command]["Action"]
		command_expected = self.commands_data[selected_command]["Expected Result"]

		# 🔹 Identificăm placeholder-ele
		required_categories = set(
			part.strip("{}") for part in command_action.split() if "{" in part
		) | set(
			part.strip("{}") for part in command_expected.split() if "{" in part
		)

		# 🔹 Deschidem dialogul pentru selecția parametrilor
		parameter_dialog = PreviewTestStepDialog(command_action, command_expected, self.parameters_data,
		                                         required_categories)

		if not parameter_dialog.exec_():
			print("❌ Parameter selection was canceled. Exiting...")
			return

		selected_parameters = parameter_dialog.selected_parameters
		print(f"✅ Selected parameters: {selected_parameters}")  # Debugging

		if not selected_parameters:
			QMessageBox.warning(self, "No Parameters Selected", "You must select parameters before applying.")
			return

		# 🔹 Înlocuim placeholder-ele cu valorile selectate
		test_data_description = []
		for category, value in selected_parameters.items():
			command_action = command_action.replace(f"{{{category}}}", value)
			command_expected = command_expected.replace(f"{{{category}}}", value)
			test_data_description.append(f"{category} = {value}")  # ✅ Stocăm ca listă de string-uri

		# 🔹 Asigurăm structura JSON corectă
		if test_name not in self.tests_data:

			return
		# 🔹 Verificăm dacă Action și Expected Results sunt liste (conform `example_test.json`)
		if not isinstance(self.tests_data[test_name].get("Action"), list):
			self.tests_data[test_name]["Action"] = []
		if not isinstance(self.tests_data[test_name].get("Expected Results"), list):
			self.tests_data[test_name]["Expected Results"] = []

		# 🔹 Adăugăm noul test step în listă
		self.tests_data[test_name]["Action"].append(command_action)
		self.tests_data[test_name]["Expected Results"].append(command_expected)

		self.update_test_after_step_edit(test_name)
		# 🔹 Apelăm funcția de actualizare a `Test Data Description`
		self.update_test_data_description(test_name)

		# 🔹 Apelăm funcția de actualizare după adăugarea unui step nou
		self.update_description_tcg(test_name)
		self.update_test_in_ui(test_name)

	def delete_test_step(self, test_name, step_index):
		"""Șterge test step-ul selectat din test."""
		if test_name not in self.tests_data:
			print(f"❌ ERROR: Test '{test_name}' not found!")
			return

		reply = QMessageBox.question(
			self, "Delete Step",
			f"Are you sure you want to delete step {step_index + 1} from '{test_name}'?",
			QMessageBox.Yes | QMessageBox.No, QMessageBox.No)

		if reply == QMessageBox.Yes:
			print(f"🗑️ Deleting step {step_index} from test {test_name}")

			# 🔹 Ștergem step-ul din toate categoriile relevante
			del self.tests_data[test_name]["Action"][step_index]
			del self.tests_data[test_name]["Expected Results"][step_index]

			self.update_test_after_step_edit(test_name)
			# 🔹 Apelăm funcția de actualizare a `Test Data Description`
			self.update_test_data_description(test_name)

			# 🔹 Apelăm funcția de actualizare după adăugarea unui step nou
			self.update_description_tcg(test_name)

	def delete_test(self, test_name):
		"""Șterge testul selectat din listă și actualizează UI-ul."""

		try:
			if test_name not in self.tests_data:
				print(f"❌ ERROR: Test '{test_name}' not found!")
				return

			reply = QMessageBox.question(
				self, "Delete Test",
				f"Are you sure you want to delete test '{test_name}'?",
				QMessageBox.Yes | QMessageBox.No, QMessageBox.No
			)

			if reply == QMessageBox.Yes:
				print(f"🗑 Deleting test {test_name}")

				del self.tests_data[test_name]  # ✅ Ștergem testul din JSON
				self.save_tests()

				# 🔹 Verificăm dacă UI-ul poate gestiona ștergerea
				if self.test_table is None:
					print("❌ ERROR: self.test_table is None. UI update failed!")
					return

				# 🔹 Căutăm rândul în UI și îl ștergem
				row_to_delete = -1
				for row in range(self.test_table.rowCount()):
					if self.test_table.item(row, 0).text().strip() == test_name:
						row_to_delete = row
						break

				if row_to_delete != -1:
					self.test_table.removeRow(row_to_delete)
					print(f"✅ Test '{test_name}' deleted from UI.")

		except Exception as e:
			print(f"❌ CRITICAL ERROR in `delete_test`: {e}")
			traceback.print_exc()

	def move_test_step(self, test_name, step_index, direction):
		"""Mută test step-ul selectat în sus sau în jos."""
		if test_name not in self.tests_data:
			print(f"❌ ERROR: Test '{test_name}' not found!")
			return

		steps = self.tests_data[test_name]["Action"]

		if direction == "up" and step_index > 0:
			print(f"🔼 Moving step {step_index} UP in test {test_name}")
			steps[step_index], steps[step_index - 1] = steps[step_index - 1], steps[step_index]
			self.tests_data[test_name]["Expected Results"][step_index], self.tests_data[test_name]["Expected Results"][
				step_index - 1] = \
				self.tests_data[test_name]["Expected Results"][step_index - 1], \
				self.tests_data[test_name]["Expected Results"][step_index]
		elif direction == "down" and step_index < len(steps) - 1:
			print(f"🔽 Moving step {step_index} DOWN in test {test_name}")
			steps[step_index], steps[step_index + 1] = steps[step_index + 1], steps[step_index]
			self.tests_data[test_name]["Expected Results"][step_index], self.tests_data[test_name]["Expected Results"][
				step_index + 1] = \
				self.tests_data[test_name]["Expected Results"][step_index + 1], \
				self.tests_data[test_name]["Expected Results"][step_index]
		else:
			print("❌ ERROR: Move not possible")
			return

		# 🔹 Apelăm funcția de actualizare după adăugarea unui step nou
		self.update_test_after_step_edit(test_name)
		self.update_description_tcg(test_name)

	def copy_test_step(self, test_name, step_index):
		"""Copiază test step-ul selectat într-un buffer temporar."""
		if test_name not in self.tests_data:
			print(f"❌ ERROR: Test '{test_name}' not found!")
			return

		self.copied_step = self.tests_data[test_name]["Action"][step_index]
		self.copied_expected = self.tests_data[test_name]["Expected Results"][step_index]

		print(f"📋 Copied step {step_index} from test: {test_name}")

	def paste_test_step(self, test_name, step_index):
		"""Lipește test step-ul copiat într-un test selectat."""
		if not self.copied_step:
			QMessageBox.warning(self, "No Copied Step", "Please copy a test step first.")
			return

		if test_name not in self.tests_data:
			print(f"❌ ERROR: Test '{test_name}' not found!")
			return

		print(f"📌 Pasting copied step into test: {test_name} at position {step_index}")

		# 🔹 Adăugăm step-ul copiat în testul selectat
		self.tests_data[test_name]["Action"].insert(step_index, self.copied_step)
		self.tests_data[test_name]["Expected Results"].insert(step_index, self.copied_expected)

		# 🔹 Apelăm funcția de actualizare a `Test Data Description`
		self.update_test_data_description(test_name)

		# 🔹 Apelăm funcția de actualizare după adăugarea unui step nou
		self.update_description_tcg(test_name)

		self.update_test_after_step_edit(test_name)

	def add_test_to_ui(self, test_name, test_data, after_test_name=None):
		"""Adaugă un test în UI imediat după un test existent."""

		if not self.test_table:
			print("❌ ERROR: test_table is None! Cannot add test to UI.")
			return

		# 🔹 Găsim poziția testului original
		row_position = self.test_table.rowCount()  # Default: ultimul rând
		if after_test_name:
			for row in range(self.test_table.rowCount()):
				if self.test_table.item(row, 0) and self.test_table.item(row, 0).text().strip() == after_test_name:
					row_position = row + 1
					break

		self.test_table.insertRow(row_position)  # ✅ Inserăm testul duplicat imediat după original

		print(f"✅ Adding test '{test_name}' at row {row_position}, after '{after_test_name}'.")

		# 🔹 `Test Name` devine read-only
		test_name_item = QTableWidgetItem(test_name)
		test_name_item.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable)  # ❌ User-ul NU poate edita acest câmp
		self.test_table.setItem(row_position, 0, test_name_item)

		# 🔹 `Description` și `Precondition` sunt editabile
		for col, field in enumerate(["Description", "Precondition"], start=1):
			item = QTableWidgetItem(test_data.get(field, ""))
			item.setFlags(
				Qt.ItemIsEnabled | Qt.ItemIsSelectable | Qt.ItemIsEditable)  # ✅ User-ul poate edita aceste celule
			self.test_table.setItem(row_position, col, item)

		# 🔹 Creăm tabele embedate pentru Action și Expected Results
		action_table = TestStepTable(test_data.get("Action", []), self.test_table, row_position)
		expected_table = TestStepTable(test_data.get("Expected Results", []), self.test_table, row_position)

		self.test_table.setCellWidget(row_position, 3, action_table)
		self.test_table.setCellWidget(row_position, 4, expected_table)

		# 🔹 Setăm `Test Data Description` și `Description for TCG` ca read-only
		for col, field in enumerate(["Test Data Description", "Description TCG"], start=5):
			item = QTableWidgetItem("\n".join(test_data.get(field, [])))
			item.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable)  # ❌ User-ul NU poate edita aceste celule
			self.test_table.setItem(row_position, col, item)

		print(f"✅ Test '{test_name}' added to UI at row {row_position}.")

	def load_tests(self):
		"""Încarcă testele din JSON la pornirea aplicației și setează UI-ul corect."""
		if hasattr(self, "initial_load_done") and self.initial_load_done:
			print("🔹 Skipping full reload - load_tests() is only called on startup.")
			return  # ✅ Prevenim apelurile după inițializare

		print("🔹 Loading tests from JSON...")

		self.test_table.blockSignals(True)  # ✅ Dezactivăm semnalele pentru a preveni salvările inutile
		self.test_table.setRowCount(0)  # Resetăm tabelul

		for test_name, test_data in self.tests_data.items():
			self.add_test_to_ui(test_name, test_data)  # ✅ Mutăm logica de încărcare per test într-o funcție separată

		self.test_table.blockSignals(False)  # ✅ Reactivăm semnalele după încărcare

		self.test_table.resizeRowsToContents()
		self.test_table.verticalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
		self.test_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)

		self.initial_load_done = True  # ✅ Marcăm că încărcarea inițială a fost efectuată
		print("✅ Tests loaded successfully.")

	def load_parameters(self):
		"""Încarcă parametrii din parameters.json."""
		if not os.path.exists(self.parameters_file):
			print("❌ Parameters file not found.")
			self.parameters_data = {}
			return

		try:
			with open(self.parameters_file, "r", encoding="utf-8") as file:
				self.parameters_data = json.load(file)
			print("✅ Parameters loaded successfully:", self.parameters_data.keys())
		except json.JSONDecodeError:
			print("❌ Error loading parameters file.")
			self.parameters_data = {}

	def load_commands(self):
		"""Încarcă comenzile din `generic_commands.json`."""
		if not os.path.exists(self.commands_file):
			print("❌ Commands file not found.")
			self.commands_data = {}
			return

		try:
			with open(self.commands_file, "r", encoding="utf-8") as file:
				self.commands_data = json.load(file)
			print("✅ Commands loaded successfully:", self.commands_data.keys())
		except json.JSONDecodeError:
			print("❌ Error loading commands file.")
			self.commands_data = {}

	def load_json(self, file_path):
		"""Încarcă un JSON sau returnează un dicționar gol dacă nu există."""
		if os.path.exists(file_path):
			try:
				with open(file_path, "r", encoding="utf-8") as file:
					return json.load(file)
			except json.JSONDecodeError:
				print(f"❌ Error: JSON file '{file_path}' is corrupted. Returning empty dictionary.")
				return {}
		return {}

	def import_from_xlsx(self):
		"""Importă testele dintr-un fișier XLSX și le adaugă doar dacă nu sunt deja în tabel și JSON."""
		file_path, _ = QFileDialog.getOpenFileName(self, "Select Excel File", "", "Excel Files (*.xlsx);;All Files (*)")
		if not file_path:
			return

		df = pd.read_excel(file_path, dtype=str).fillna("")

		for _, row in df.iterrows():
			test_name = row["Test Name"].strip()

			# Verificăm dacă testul există deja în JSON sau UI
			if test_name in self.tests_data or self.is_test_in_table(test_name):
				print(f"⚠️ Test '{test_name}' already exists. Skipping import.")
				continue  # Sărim peste acest test, deoarece este deja prezent

			# Adăugăm testul în JSON și UI
			self.tests_data[test_name] = row.to_dict()
			self.add_test_to_table(test_name, self.tests_data[test_name])

		self.save_tests()
		QMessageBox.information(self, "Import Completed", "Tests imported successfully from XLSX!")

	def export_to_xlsx(self):
		file_path, _ = QFileDialog.getSaveFileName(self, "Save Excel File", "", "Excel Files (*.xlsx);;All Files (*)")
		if not file_path:
			return

		wb = Workbook()
		ws = wb.active
		ws.title = "Tests"

		headers = ["Test Name", "Description", "Precondition", "Action", "Expected Results", "Test Data Description",
		           "Description TCG"]
		ws.append(headers)

		blue_fill = PatternFill(start_color="007BFF", end_color="007BFF", fill_type="solid")
		bold_font = Font(bold=True, color="FFFFFF")

		for col_num, header in enumerate(headers, 1):
			cell = ws.cell(row=1, column=col_num, value=header)
			cell.fill = blue_fill
			cell.font = bold_font
			cell.alignment = Alignment(horizontal="center", vertical="center")

		for test_name, test_data in self.tests_data.items():
			action_steps = test_data.get("Steps", [])
			action_text = "\n".join([f"{i + 1}. {step['Action']}" for i, step in enumerate(action_steps)])
			expected_text = "\n".join([f"{i + 1}. {step['Expected']}" for i, step in enumerate(action_steps)])
			test_data_description = "\n".join(test_data.get("Test Data Description", "").split(", "))
			description_tcg = "\n".join(test_data.get("Description TCG", "").split(", "))

			row = [
				test_name,
				test_data.get("Description", ""),
				test_data.get("Precondition", ""),
				action_text,
				expected_text,
				test_data_description,
				description_tcg
			]
			ws.append(row)

		for col_num in range(1, 4):
			for row_num in range(2, ws.max_row + 1):
				ws.cell(row=row_num, column=col_num).alignment = Alignment(horizontal="center", vertical="center")

		wb.save(file_path)
		QMessageBox.information(self, "Export Completed", "Tests exported successfully to XLSX!")

	def save_tests(self):
		"""Salvează testele în JSON."""
		try:
			with open(self.json_file, "w", encoding="utf-8") as file:
				json.dump(self.tests_data, file, indent=4)
			print("✅ Tests saved successfully.")
		except Exception as e:
			print(f"❌ Error saving tests: {e}")

	def get_selected_test_step(self, selected_row, selected_col):
		"""Returnează testul și index-ul step-ului selectat din Action sau Expected Results."""

		print(f"🔹 DEBUG: Checking test step at Row: {selected_row}, Column: {selected_col}")

		if selected_row == -1 or selected_col not in [3, 4]:  # ✅ Asigurăm că e doar pe Action sau Expected
			print("❌ No step selected - Wrong column")
			return None, None, None

		test_name = self.test_table.item(selected_row, 0).text().strip()
		cell_widget = self.test_table.cellWidget(selected_row, selected_col)

		if not test_name or not isinstance(cell_widget, TestStepTable):
			print("❌ No step table detected in the selected cell")
			return None, None, None

		step_index = cell_widget.table.currentRow()

		print(f"🔹 DEBUG: Step Index in Table: {step_index}")

		if step_index == -1:
			print("❌ No valid step selected inside the embedded table")
			return None, None, None

		print(f"✅ Step Selected: {step_index} in {test_name} (Column: {selected_col})")

		return test_name, step_index, selected_col

	def show_context_menu(self, position):
		"""Afișează un meniu contextual doar dacă utilizatorul a făcut click dreapta pe Action sau Expected Results."""

		index = self.test_table.indexAt(position)  # ✅ Detectează exact unde s-a făcut click
		selected_row = index.row()
		selected_col = index.column()

		print(f"🔹 DEBUG: Clicked Row: {selected_row}, Clicked Column: {selected_col}")

		if selected_row == -1 or selected_col not in [3, 4]:  # ✅ Asigurăm că e doar pe Action sau Expected
			print("❌ DEBUG: Click dreapta pe o coloană greșită")
			return

		test_name, step_index, selected_col = self.get_selected_test_step(selected_row, selected_col)

		if test_name is None or step_index is None:
			print("❌ DEBUG: No valid test step selected for context menu")
			return

		menu = QMenu(self)

		delete_action = QAction("❌ Delete Step", self)
		delete_action.triggered.connect(lambda: self.delete_test_step(test_name, step_index))
		menu.addAction(delete_action)

		move_up_action = QAction("🔼 Move Step Up", self)
		move_up_action.triggered.connect(lambda: self.move_test_step(test_name, step_index, "up"))
		menu.addAction(move_up_action)

		move_down_action = QAction("🔽 Move Step Down", self)
		move_down_action.triggered.connect(lambda: self.move_test_step(test_name, step_index, "down"))
		menu.addAction(move_down_action)

		copy_action = QAction("📋 Copy Step", self)
		copy_action.triggered.connect(lambda: self.copy_test_step(test_name, step_index))
		menu.addAction(copy_action)

		paste_action = QAction("📌 Paste Step", self)
		paste_action.triggered.connect(lambda: self.paste_test_step(test_name, step_index))
		menu.addAction(paste_action)

		print("✅ DEBUG: Context menu displayed")
		menu.exec_(self.test_table.viewport().mapToGlobal(position))

	def show_test_context_menu(self, position):
		"""Afișează un meniu contextual pentru teste doar dacă utilizatorul nu a selectat un test step."""

		# ✅ Dacă tocmai s-a deschis meniul pentru un test step, prevenim deschiderea meniului pentru teste
		if getattr(self, "suppress_test_context_menu", False):
			self.suppress_test_context_menu = False  # ✅ Resetăm flag-ul
			print("⚠️ DEBUG: Suppressed test context menu")
			return

		index = self.test_table.indexAt(position)  # ✅ Detectează unde s-a făcut click
		selected_row = index.row()

		if selected_row == -1:
			print("❌ DEBUG: No valid test selected for context menu")
			return

		test_name = self.test_table.item(selected_row, 0).text().strip()
		print(f"🔹 DEBUG: Clicked on test: {test_name}")

		menu = QMenu(self)

		# 🔹 Ștergere Test
		delete_action = QAction("🗑 Delete Test", self)
		delete_action.triggered.connect(lambda: self.run_and_close_menu(self.delete_test, test_name, menu))
		menu.addAction(delete_action)

		# 🔹 Adăugare Test Step
		add_step_action = QAction("➕ Add Test Step", self)
		add_step_action.triggered.connect(
			lambda: self.run_and_close_menu(self.add_test_step_from_menu, test_name, menu))
		menu.addAction(add_step_action)

		# 🔹 Actualizare Test Data Description
		update_test_data_action = QAction("🔄 Update Test Data", self)
		update_test_data_action.triggered.connect(
			lambda: self.run_and_close_menu(self.update_test_data_description, test_name, menu))
		menu.addAction(update_test_data_action)

		# 🔹 Actualizare Description for TCG
		update_tcg_action = QAction("🔄 Update Description for TCG", self)
		update_tcg_action.triggered.connect(
			lambda: self.run_and_close_menu(self.update_description_tcg, test_name, menu))
		menu.addAction(update_tcg_action)

		# 🔹 **Duplicate Test Case**
		duplicate_test_action = QAction("📑 Duplicate Test Case", self)
		duplicate_test_action.triggered.connect(
			lambda: self.run_and_close_menu(self.duplicate_test_case, test_name, menu))
		menu.addAction(duplicate_test_action)

		# 🔹 **Edit Test Name**
		edit_test_name_action = QAction("✏️ Edit Test Name", self)
		edit_test_name_action.triggered.connect(lambda: self.run_and_close_menu(self.edit_test_name, test_name, menu))
		menu.addAction(edit_test_name_action)

		print("✅ DEBUG: Context menu displayed")
		menu.exec_(self.test_table.viewport().mapToGlobal(position))

	def add_test_step_from_menu(self, test_name):
		"""Adaugă un nou Test Step pentru testul selectat din meniul contextual."""
		print(f"➕ Adding test step to: {test_name}")

		# 🔹 Simulăm selecția rândului pentru ca `add_test_step()` să funcționeze corect
		for row in range(self.test_table.rowCount()):
			if self.test_table.item(row, 0).text().strip() == test_name:
				self.test_table.selectRow(row)
				break

		self.add_test_step()

	def update_test_data_description(self, test_name):
		if test_name not in self.tests_data:
			print(f"❌ ERROR: Test '{test_name}' not found!")
			return

		print(f"🔹 Updating Test Data Description for: {test_name}")

		variant_groups = {}

		# Iterăm peste câmpurile Action și Expected Results
		for field in ["Action", "Expected Results"]:
			steps = self.tests_data[test_name].get(field, [])
			for step in steps:
				words = step.split()
				for word in words:
					# Verificăm existența parametrului în parameters.json
					for category, param_dict in self.parameters_data.items():
						if word in param_dict:
							value_dict = param_dict[word]
							for variant, value in value_dict.items():
								if value != "":
									variant_groups = variant_groups if 'variant_groups' in locals() else {}
									variant_groups.setdefault(variant, set()).add(f"{word} = {value}")

		# Construim structura Test Data Description în format listă
		new_test_data_description = []
		for variant, values in variant_groups.items():
			new_test_data_description.append(f"{variant}:")
			new_test_data_description.extend(sorted(values))
			new_test_data_description.append("")  # Adaugăm spațiu după fiecare variantă

		# Verificăm dacă există schimbări față de versiunea actuală
		current_data_description = self.tests_data[test_name].get("Test Data Description", [])
		if new_test_data_description == current_data_description:
			print(f"✅ No changes detected for {test_name}, skipping update.")
			return

		# Actualizăm datele în JSON și UI
		self.tests_data[test_name]["Test Data Description"] = new_test_data_description
		self.save_tests()
		self.update_test_in_ui(test_name)

		print(f"✅ Updated Test Data Description for {test_name}.")

	def update_description_tcg(self, test_name):
		"""Generează automat Description for TCG doar dacă există o modificare."""

		if test_name not in self.tests_data:
			print(f"❌ ERROR: Test '{test_name}' not found!")
			return

		print(f"🔹 Checking if Description for TCG needs an update for: {test_name}")

		actions = self.tests_data[test_name].get("Action", [])
		expected_results = self.tests_data[test_name].get("Expected Results", [])

		if not actions or not expected_results:
			if self.tests_data[test_name].get("Description TCG", []):  # ✅ Doar dacă nu este deja gol
				self.tests_data[test_name]["Description TCG"] = []
				self.save_tests()
				self.update_test_in_ui(test_name)  # ✅ Actualizăm doar testul modificat
				print(f"✅ Cleared Description for TCG for {test_name}.")
			return

		description_tcg = ["PRECONDITION:"]

		precondition_text = self.tests_data[test_name].get("Precondition", "").strip()
		if precondition_text and "call" not in precondition_text.lower():
			description_tcg.append(f"1. {precondition_text}")
			description_tcg.append("")

		description_tcg.append("ACTION:")

		for index, (action, expected) in enumerate(zip(actions, expected_results), start=1):
			formatted_action = self.format_step(action)
			formatted_expected = self.format_step(expected)
			description_tcg.append(f"{index}. {formatted_action} {formatted_expected}")

		# ✅ Verificăm dacă `Description for TCG` există și este diferit
		if self.tests_data[test_name].get("Description TCG", []) == description_tcg:
			print(f"✅ No changes detected in Description for TCG for {test_name}, skipping update.")
			return

		# 🔹 Actualizăm JSON și UI doar dacă există modificări
		self.tests_data[test_name]["Description TCG"] = description_tcg
		self.save_tests()
		self.update_test_in_ui(test_name)  # ✅ Actualizăm doar testul modificat în UI

		print(f"✅ Updated Description for TCG for {test_name}: {description_tcg}")

	def format_step(self, step_text):
		"""Înlocuiește parametrii dintr-un test step cu versiunea lor încadrată între ' '."""
		words = step_text.split()
		formatted_words = []

		for word in words:
			# Verificăm dacă acest cuvânt este un parametru valid în `parameters.json`
			for category, param_dict in self.parameters_data.items():
				if word in param_dict:  # ✅ Parametru detectat, adăugăm ' '
					word = f"'{word}'"
					break

			formatted_words.append(word)

		return " ".join(formatted_words)  # ✅ Reconstruim fraza cu parametrii formatați

	def save_edited_test(self, row, column):
		"""Salvează automat modificările făcute de user în `Description` și `Precondition` și actualizează doar acel test în UI."""

		if column not in [1, 2]:  # ✅ Edităm doar `Description` și `Precondition`
			return

		item = self.test_table.item(row, column)
		if item is None:
			return

		test_name = self.test_table.item(row, 0).text().strip()

		if test_name not in self.tests_data:
			print(f"❌ ERROR: Test '{test_name}' not found in JSON!")
			return

		field_map = {1: "Description", 2: "Precondition"}
		field_name = field_map[column]
		new_value = item.text().strip()

		print(f"📝 Updating '{field_name}' for '{test_name}' with: {new_value}")

		self.tests_data[test_name][field_name] = new_value
		self.save_tests()

		# 🔹 Actualizăm doar testul modificat în UI
		self.update_test_in_ui(test_name)

	def duplicate_test_case(self, test_name):
		"""Creează o copie a unui test și îl salvează în JSON și UI imediat după original."""

		try:
			if test_name not in self.tests_data:
				print(f"❌ ERROR: Test '{test_name}' not found!")
				return

			print(f"📑 Duplicating test: {test_name}")

			# 🔹 Creăm un nou nume de test incrementat automat
			base_name = f"{test_name}_"
			count = 1

			while f"{base_name}{count}" in self.tests_data:  # ✅ Verificăm dacă există deja
				count += 1

			new_test_name = f"{base_name}{count}"
			print(f"🔹 New test name generated: {new_test_name}")

			# 🔹 Copiem testul original folosind `deepcopy()` pentru a evita referințele comune
			new_test_data = copy.deepcopy(self.tests_data[test_name])
			print(f"✅ Test '{test_name}' copied successfully.")

			# 🔹 Adăugăm noul test în JSON imediat după original
			test_names = list(self.tests_data.keys())  # ✅ Păstrăm ordinea testelor
			index = test_names.index(test_name) + 1
			new_tests_data = {k: self.tests_data[k] for k in test_names[:index]}  # ✅ Păstrăm testele de până la index
			new_tests_data[new_test_name] = new_test_data  # ✅ Adăugăm testul duplicat
			new_tests_data.update({k: self.tests_data[k] for k in test_names[index:]})  # ✅ Păstrăm restul testelor

			self.tests_data = new_tests_data
			print(f"✅ Test '{new_test_name}' added to tests_data after '{test_name}'.")

			# 🔹 Salvăm modificările
			self.save_tests()

			# 🔹 Adăugăm testul în UI imediat după original
			self.add_test_to_ui(new_test_name, new_test_data, after_test_name=test_name)

			# 🔹 Așteptăm 100ms înainte de actualizare pentru a preveni crash-ul
			QTimer.singleShot(100, lambda: self.update_test_in_ui(new_test_name))

			print(f"✅ Test '{test_name}' duplicated as '{new_test_name}'")

		except Exception as e:
			print(f"❌ CRITICAL ERROR in `duplicate_test_case`: {e}")
			import traceback
			traceback.print_exc()

	def edit_test_name(self, old_test_name):
		"""Permite user-ului să editeze numele unui test și actualizează doar rândul din UI."""
		if old_test_name not in self.tests_data:
			return
		# 🔹 Cerem user-ului un nou nume
		new_test_name, ok = QInputDialog.getText(self, "Edit Test Name", "Enter new test name:", QLineEdit.Normal,
		                                         old_test_name)
		if not ok or not new_test_name.strip():  # ✅ User-ul a apăsat cancel sau nu a introdus nimic
			print("❌ Edit Test Name canceled.")
			return

		new_test_name = new_test_name.strip()

		# 🔹 Verificăm dacă numele deja există
		if new_test_name in self.tests_data:
			QMessageBox.warning(self, "Test Name Exists",
			                    f"A test named '{new_test_name}' already exists. Please choose another name.")
			return

		# 🔹 Salvăm poziția rândului
		row_position = -1
		for row in range(self.test_table.rowCount()):
			if self.test_table.item(row, 0).text().strip() == old_test_name:
				row_position = row
				break

		if row_position == -1:
			print(f"❌ ERROR: Test '{old_test_name}' not found in table!")
			return

		# 🔹 Redenumim testul în JSON și păstrăm toate datele
		self.tests_data[new_test_name] = self.tests_data.pop(old_test_name)

		# 🔹 Salvăm modificările
		self.save_tests()

		# 🔹 Actualizăm doar rândul în UI
		self.test_table.item(row_position, 0).setText(new_test_name)

	def run_and_close_menu(self, function, test_name, menu):
		"""Execută funcția selectată și închide meniul de click dreapta."""
		function(test_name)  # ✅ Executăm funcția selectată
		menu.close()  # ✅ Închidem meniul după execuție

	def update_test_after_step_edit(self, test_name):
		"""Actualizează doar testul modificat în UI după modificarea unui test step."""

		# 🔹 Salvăm testele și actualizăm doar testul modificat
		self.save_tests()
		self.update_test_in_ui(test_name)

		print(f"✅ UI updated after test step modification for '{test_name}'.")

	def update_test_in_ui(self, test_name):
		"""Actualizează un test existent în UI fără să reîncarce toată lista."""

		try:
			print(f"🔄 Updating test '{test_name}' in UI...")

			row_position = -1
			for row in range(self.test_table.rowCount()):
				if self.test_table.item(row, 0) and self.test_table.item(row, 0).text().strip() == test_name:
					row_position = row
					break

			if row_position == -1:
				print(f"❌ ERROR: Test '{test_name}' not found in UI! Skipping update.")
				return

			# 🔹 Verificăm dacă testul există în JSON înainte de update
			if test_name not in self.tests_data:
				print(f"❌ ERROR: Test '{test_name}' not found in JSON after step addition! Skipping UI update.")
				return

			print(f"✅ Test '{test_name}' found at row {row_position}, updating UI...")

			# 🔹 Actualizăm `Description` și `Precondition`
			for col, field in enumerate(["Description", "Precondition"], start=1):
				if self.test_table.item(row_position, col):
					self.test_table.item(row_position, col).setText(self.tests_data[test_name].get(field, ""))
				else:
					print(f"⚠️ WARNING: Column {col} is None for test '{test_name}', skipping update.")

			# 🔹 Actualizăm `Test Data Description` și `Description for TCG`
			for col, field in enumerate(["Test Data Description", "Description TCG"], start=5):
				if self.test_table.item(row_position, col):
					self.test_table.item(row_position, col).setText(
						"\n".join(self.tests_data[test_name].get(field, [])))
				else:
					print(f"⚠️ WARNING: Column {col} is None for test '{test_name}', skipping update.")

			# 🔹 Reîmprospătăm `Action` și `Expected Results`
			action_table = TestStepTable(self.tests_data[test_name].get("Action", []), self.test_table, row_position)
			expected_table = TestStepTable(self.tests_data[test_name].get("Expected Results", []), self.test_table,
			                               row_position)

			self.test_table.setCellWidget(row_position, 3, action_table)
			self.test_table.setCellWidget(row_position, 4, expected_table)

			print(f"✅ Test '{test_name}' updated in UI at row {row_position}.")

		except Exception as e:
			print(f"❌ CRITICAL ERROR in `update_test_in_ui`: {e}")
			import traceback
			traceback.print_exc()

	def filter_tests(self, text, table):
		"""Filters tests based on search input."""
		text = text.strip().lower()  # Convert search text to lowercase

		for row in range(table.rowCount()):
			test_name_item = table.item(row, 0)  # Get the test name
			if test_name_item:
				test_name = test_name_item.text().strip().lower()
				table.setRowHidden(row, text not in test_name)  # ✅ Hide rows that don't match

	def get_resource_path(self, relative_path):
		"""Get the correct path whether running as a script or an executable."""
		if getattr(sys, 'frozen', False):  # Running as compiled .exe
			base_path = os.path.dirname(sys.executable) # Use folder where main.exe is
		else:
			base_path = os.path.abspath(os.path.dirname(__file__))  # Development mode

		return os.path.join(base_path, relative_path)


















